import openpyxl
import os
from .utils import URL_PATTERN, clean_urls

def extract_urls_from_excel(filepath):
    urls = []
    workbook = None
    try:
        workbook = openpyxl.load_workbook(filepath, data_only=True) #data_only=True gets cell values instead of formulas
        #Iterate through all sheets in the workbook
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            #Iterate through all rows and cells in the sheet
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.hyperlink and cell.hyperlink.target: # check for hyperlinks in the worksheet
                        urls.append(cell.hyperlink.target)


    except Exception as e:
        print(f"--> Error processing Excel file '{os.path.basename(filepath)}': {e}")
        return []
    finally:
        if workbook:
             pass
    cleaned_list = clean_urls(urls)
    return cleaned_list